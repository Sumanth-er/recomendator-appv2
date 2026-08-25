"""Historical purchase prices - the SAP BW extract (spec 2.2).

The extract arrives as a workbook with two sheets:

  PO Price History  one row per purchase order line
  Price Summary     average / min / max / last invoiced price per material

Sheet 2 is the price benchmark section 4.2 asks for. Sheet 1 adds what sheet 2
cannot: which vendors the spend actually went to, which turns the dual-sourcing
concentration check from an assumed split into a measured one.

Both are matched to quote lines on CAS number, never on material description or
material number. A row whose CAS falls outside the category is reported rather
than absorbed.

A single-sheet CSV of either shape is also accepted.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import HistoricalPrice, HistoricalPurchase, Material
from .pipeline import supplier_key

log = logging.getLogger(__name__)

CAS_RE = re.compile(r"\b\d{2,7}-\d{2}-\d\b")

# --- column names, as they appear in the extract ---------------------------

SUMMARY_ALIASES = {
    "cas_no": ("cas number", "cas no", "cas", "casnr"),
    "material_number": ("material number", "material no", "matnr", "material"),
    "material_description": ("material description", "description"),
    "avg_price_eur_l": ("avg net price eur/l", "average net price eur/l",
                        "avg price eur/l", "average price eur/l", "avg price"),
    "min_price_eur_l": ("min net price eur/l", "minimum net price eur/l",
                        "min price eur/l", "min price"),
    "max_price_eur_l": ("max net price eur/l", "maximum net price eur/l",
                        "max price eur/l", "max price"),
    "last_invoiced_price_eur_l": ("last invoice price eur/l",
                                  "last invoiced price eur/l",
                                  "last invoice price", "last invoiced price"),
    "po_line_count": ("number of po line items", "po line items", "line items",
                      "number of lines"),
}

HISTORY_ALIASES = {
    "cas_no": ("cas number", "cas no", "cas", "casnr"),
    "material_number": ("material number", "material no", "matnr"),
    "material_description": ("material description", "description"),
    "category": ("category",),
    "plant": ("plant",),
    "plant_name": ("plant name",),
    "vendor_number": ("vendor number", "vendor no", "lifnr"),
    "vendor_name": ("vendor name", "vendor", "supplier name"),
    "po_number": ("po number", "purchase order", "ebeln"),
    "po_item": ("po item", "item"),
    "document_date": ("document date", "doc date", "po date"),
    "delivery_date": ("delivery date",),
    "order_quantity": ("order quantity", "quantity", "menge"),
    "base_uom": ("base uom", "uom", "unit of measure"),
    "net_price": ("net price",),
    "price_unit": ("price unit",),
    "currency": ("currency", "curr"),
    "net_value_cur": ("net value cur", "net value currency"),
    "net_value_eur": ("net value eur", "net value in eur"),
    "purchasing_group": ("purchase group", "purchasing group"),
    "incoterm": ("incoterm", "incoterms"),
    "unit_price_eur_l": ("unit price eur/l", "unit price"),
}

SUMMARY_SHEET_HINTS = ("price summary", "summary")
HISTORY_SHEET_HINTS = ("po price history", "price history", "history", "po")


def _normalise(header) -> str:
    text = re.sub(r"[^a-z0-9/]+", " ", str(header or "").lower())
    return re.sub(r"\s+", " ", text).strip()


def _map_headers(headers, aliases: dict) -> dict[str, int]:
    """Field name -> column index, for whichever columns the sheet has."""
    cleaned = {}
    for index, header in enumerate(headers):
        key = _normalise(header)
        if key and key not in cleaned:
            cleaned[key] = index

    mapping: dict[str, int] = {}
    for field, options in aliases.items():
        for option in options:
            if option in cleaned:
                mapping[field] = cleaned[option]
                break
    return mapping


def _dec(value) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    text = re.sub(r"[^\d.,-]", "", str(value))
    if not text:
        return None
    # A comma before one or two trailing digits is a decimal separator.
    if re.search(r",\d{1,2}$", text) and "." not in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _int(value) -> int | None:
    d = _dec(value)
    return int(d) if d is not None else None


def _date(value) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _cas(value) -> str | None:
    match = CAS_RE.search(str(value or ""))
    return match.group(0) if match else None


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def _load_summary(session: Session, rows, filename: str, known: set[str]) -> dict:
    """Sheet 2 - the per-material price benchmark (spec 2.2)."""
    if not rows:
        return {"loaded": 0, "skipped": 0, "errors": ["price summary sheet is empty"]}

    mapping = _map_headers(rows[0], SUMMARY_ALIASES)
    if "cas_no" not in mapping:
        return {"loaded": 0, "skipped": 0,
                "errors": ["no CAS number column in the price summary sheet"]}

    loaded, skipped, errors = 0, 0, []
    for number, row in enumerate(rows[1:], start=2):
        get = lambda f: row[mapping[f]] if f in mapping and mapping[f] < len(row) else None  # noqa: E731
        cas = _cas(get("cas_no"))
        if not cas:
            skipped += 1
            continue
        if cas not in known:
            skipped += 1
            errors.append(f"summary row {number}: CAS {cas} is not a material in this category")
            continue

        record = session.get(HistoricalPrice, cas) or HistoricalPrice(cas_no=cas)
        record.material_number = str(get("material_number") or "").strip() or None
        record.material_description = str(get("material_description") or "").strip() or None
        record.avg_price_eur_l = _dec(get("avg_price_eur_l"))
        record.min_price_eur_l = _dec(get("min_price_eur_l"))
        record.max_price_eur_l = _dec(get("max_price_eur_l"))
        record.last_invoiced_price_eur_l = _dec(get("last_invoiced_price_eur_l"))
        record.po_line_count = _int(get("po_line_count"))
        record.source_filename = filename
        session.add(record)
        loaded += 1

    return {"loaded": loaded, "skipped": skipped, "errors": errors,
            "columns_matched": sorted(mapping)}


def _load_history(session: Session, rows, filename: str, known: set[str]) -> dict:
    """Sheet 1 - purchase order lines, for vendor spend and period."""
    if not rows:
        return {"loaded": 0, "skipped": 0, "errors": []}

    mapping = _map_headers(rows[0], HISTORY_ALIASES)
    if "cas_no" not in mapping:
        return {"loaded": 0, "skipped": 0,
                "errors": ["no CAS number column in the PO history sheet"]}

    # A re-upload replaces the history rather than doubling every vendor total.
    for existing in session.scalars(select(HistoricalPurchase)):
        session.delete(existing)
    session.flush()

    loaded, skipped, errors = 0, 0, []
    for number, row in enumerate(rows[1:], start=2):
        get = lambda f: row[mapping[f]] if f in mapping and mapping[f] < len(row) else None  # noqa: E731
        cas = _cas(get("cas_no"))
        if not cas:
            skipped += 1
            continue
        if cas not in known:
            skipped += 1
            if len(errors) < 10:
                errors.append(f"PO row {number}: CAS {cas} is outside this category")
            continue

        vendor = str(get("vendor_name") or "").strip() or None
        session.add(HistoricalPurchase(
            cas_no=cas,
            material_number=str(get("material_number") or "").strip() or None,
            material_description=str(get("material_description") or "").strip() or None,
            category=str(get("category") or "").strip() or None,
            plant=str(get("plant") or "").strip() or None,
            plant_name=str(get("plant_name") or "").strip() or None,
            vendor_number=str(get("vendor_number") or "").strip() or None,
            vendor_name=vendor,
            vendor_key=supplier_key(vendor) if vendor else None,
            po_number=str(get("po_number") or "").strip() or None,
            po_item=str(get("po_item") or "").strip() or None,
            document_date=_date(get("document_date")),
            delivery_date=_date(get("delivery_date")),
            order_quantity=_dec(get("order_quantity")),
            base_uom=str(get("base_uom") or "").strip()[:16] or None,
            net_price=_dec(get("net_price")),
            price_unit=_dec(get("price_unit")),
            currency=str(get("currency") or "").strip()[:3] or None,
            net_value_cur=_dec(get("net_value_cur")),
            net_value_eur=_dec(get("net_value_eur")),
            purchasing_group=str(get("purchasing_group") or "").strip() or None,
            incoterm=str(get("incoterm") or "").strip()[:16] or None,
            unit_price_eur_l=_dec(get("unit_price_eur_l")),
            source_filename=filename,
        ))
        loaded += 1

    return {"loaded": loaded, "skipped": skipped, "errors": errors,
            "columns_matched": sorted(mapping)}


def _fill_summary_gaps(session: Session, filename: str) -> list[str]:
    """Derive a benchmark for any material the summary sheet did not cover.

    Computed from the PO lines and labelled, so the dashboard can say the figure
    was derived rather than taken from the summary the buyer supplied.
    """
    filled = []
    by_cas: dict[str, list[HistoricalPurchase]] = {}
    for row in session.scalars(select(HistoricalPurchase)):
        if row.cas_no and row.unit_price_eur_l is not None:
            by_cas.setdefault(row.cas_no, []).append(row)

    for cas, rows in by_cas.items():
        record = session.get(HistoricalPrice, cas)
        if record and record.avg_price_eur_l is not None:
            # Period and line count still come from the PO lines.
            dates = [r.document_date for r in rows if r.document_date]
            record.period_from = min(dates) if dates else None
            record.period_to = max(dates) if dates else None
            if record.po_line_count is None:
                record.po_line_count = len(rows)
            continue

        prices = [Decimal(str(r.unit_price_eur_l)) for r in rows]
        dated = sorted((r for r in rows if r.document_date),
                       key=lambda r: r.document_date)
        record = record or HistoricalPrice(cas_no=cas)
        record.avg_price_eur_l = (sum(prices) / len(prices)).quantize(Decimal("0.000001"))
        record.min_price_eur_l = min(prices)
        record.max_price_eur_l = max(prices)
        if dated:
            record.last_invoiced_price_eur_l = dated[-1].unit_price_eur_l
            record.last_invoiced_date = dated[-1].document_date
            record.period_from = dated[0].document_date
            record.period_to = dated[-1].document_date
        record.po_line_count = len(rows)
        record.source_filename = f"{filename} (derived from PO lines)"
        session.add(record)
        filled.append(cas)

    return filled


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _sheet_rows(worksheet) -> list[list]:
    return [list(row) for row in worksheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)]


def _pick_sheet(workbook, hints, aliases):
    """Choose a sheet by name, falling back to whichever one has the columns."""
    for name in workbook.sheetnames:
        if any(hint in _normalise(name) for hint in hints):
            return workbook[name]
    for name in workbook.sheetnames:
        rows = _sheet_rows(workbook[name])
        if rows and len(_map_headers(rows[0], aliases)) >= 3:
            return workbook[name]
    return None


def load(session: Session, content: bytes, filename: str) -> dict:
    """Load an extract. Accepts .xlsx with both sheets, or a single-sheet CSV."""
    known = {m.cas_no for m in session.scalars(select(Material))}
    name = (filename or "").lower()

    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True,
                                          data_only=True)
        try:
            history_sheet = _pick_sheet(workbook, HISTORY_SHEET_HINTS, HISTORY_ALIASES)
            summary_sheet = _pick_sheet(workbook, SUMMARY_SHEET_HINTS, SUMMARY_ALIASES)
            if history_sheet is summary_sheet:
                summary_sheet = None

            history = ({"loaded": 0, "skipped": 0, "errors": []} if history_sheet is None
                       else _load_history(session, _sheet_rows(history_sheet),
                                          filename, known))
            summary = ({"loaded": 0, "skipped": 0, "errors": []} if summary_sheet is None
                       else _load_summary(session, _sheet_rows(summary_sheet),
                                          filename, known))
        finally:
            workbook.close()
    else:
        text = content.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = [r for r in csv.reader(io.StringIO(text), dialect) if any(r)]
        # A single sheet is whichever shape its headers match.
        if rows and len(_map_headers(rows[0], HISTORY_ALIASES)) > \
                len(_map_headers(rows[0], SUMMARY_ALIASES)):
            history = _load_history(session, rows, filename, known)
            summary = {"loaded": 0, "skipped": 0, "errors": []}
        else:
            summary = _load_summary(session, rows, filename, known)
            history = {"loaded": 0, "skipped": 0, "errors": []}

    session.flush()
    derived = _fill_summary_gaps(session, filename) if history["loaded"] else []
    session.commit()

    result = {
        "summary_rows": summary["loaded"],
        "po_lines": history["loaded"],
        "skipped": summary["skipped"] + history["skipped"],
        "derived_from_po_lines": derived,
        "errors": (summary.get("errors", []) + history.get("errors", []))[:20],
        "vendors": vendor_spend_summary(session),
    }
    log.info("historical extract %s: %s summary rows, %s PO lines",
             filename, result["summary_rows"], result["po_lines"])
    return result


# ---------------------------------------------------------------------------
# Vendor spend - the measured input to the concentration check (spec 5.5)
# ---------------------------------------------------------------------------

def vendor_spend(session: Session) -> dict[str, Decimal]:
    """Historical spend in EUR per vendor key, across the PO lines."""
    totals: dict[str, Decimal] = {}
    for row in session.scalars(select(HistoricalPurchase)):
        if not row.vendor_key:
            continue
        value = row.net_value_eur
        if value is None:
            continue
        totals[row.vendor_key] = totals.get(row.vendor_key, Decimal(0)) + Decimal(str(value))
    return totals


def vendor_spend_summary(session: Session) -> list[dict]:
    """Vendor spend with each vendor's share, largest first."""
    totals = vendor_spend(session)
    grand = sum(totals.values())
    names = {}
    for row in session.scalars(select(HistoricalPurchase)):
        if row.vendor_key and row.vendor_key not in names:
            names[row.vendor_key] = row.vendor_name

    rows = [
        {
            "vendor_key": key,
            "vendor_name": names.get(key, key),
            "spend_eur": str(value.quantize(Decimal("0.01"))),
            "share_pct": (str((value / grand * 100).quantize(Decimal("0.1")))
                          if grand else None),
        }
        for key, value in totals.items()
    ]
    rows.sort(key=lambda r: Decimal(r["spend_eur"]), reverse=True)
    return rows
